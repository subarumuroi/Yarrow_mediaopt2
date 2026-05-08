"""
Generate LHS media matrix in ProgrammaticPipetting.jl input format.

Output: Excel file with:
- Main sheet: 6 compound rows x 65 columns (1_Control + 64 LHS conditions)
- Constraints sheet: compound rows with placeholder pipette/solubility values
- Miscellaneous sheet: max volume and iteration count

Compound rows:
1. KH2PO4 (macronutrient)
2. (NH4)2SO4 (macronutrient)
3. MgSO4.7H2O (macronutrient)
4. Glucose (macronutrient)
5. Trace metals (EDTA as reference, 9 secondary compounds as ratios)
6. Vitamins (myo-Inositol as reference, 6 secondary compounds as ratios)
"""

import numpy as np
import pandas as pd
from scipy.stats import qmc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Delft concentrations (g/L in final well) ──────────────────────────────────

MACRONUTRIENTS = {
    "Potassium phosphate":             13.68,
    "Ammonium sulphate":                7.13,
    "Magnesium sulphate heptahydrate":  0.475,
    "Glucose":                         20.00,
}

# Trace metals: reference compound first, then secondaries
TRACE_METALS = {
    "EDTA":                            0.030,   # reference
    "Manganese chloride tetrahydrate":  0.002,
    "Copper sulphate pentahydrate":     0.0006,
    "Calcium chloride dihydrate":       0.009,
    "Boric acid":                       0.002,
    "Zinc sulphate heptahydrate":       0.009,
    "Cobalt chloride hexahydrate":      0.0006,
    "Sodium molybdate dihydrate":       0.0008,
    "Iron sulphate heptahydrate":       0.006,
    "Potassium iodide":                 0.0002,
}

# Vitamins: reference compound first, then secondaries
VITAMINS = {
    "myo-Inositol":                    0.025,   # reference
    "Biotin":                          0.00005,
    "p-Aminobenzoic acid":             0.0002,
    "Nicotinic acid":                   0.001,
    "Calcium pantothenate":             0.001,
    "Pyridoxine HCl":                   0.001,
    "Thiamine HCl":                     0.001,
}

# ── LHS bounds (g/L in final well) ───────────────────────────────────────────

BOUNDS = {
    "Potassium phosphate":              (1.00,  27.36),
    "Ammonium sulphate":                (3.56,  14.25),
    "Magnesium sulphate heptahydrate":  (0.238,  2.50),
    "Glucose":                         (10.00, 40.00),
    "Trace metals":   (0.5 * 0.030, 2.0 * 0.030),  # 0.5x-2x of EDTA reference
    "Vitamins":       (0.5 * 0.025, 2.0 * 0.025),  # 0.5x-2x of inositol reference
}

# ── Settings ──────────────────────────────────────────────────────────────────

N_CONDITIONS = 511
SEED         = 42
TOTAL_VOL_UL = 350
MAX_ITER     = 40

# ── Helpers ───────────────────────────────────────────────────────────────────

def build_compound_name(compounds: dict) -> str:
    """
    Build compound name string in ProgrammaticPipetting format.
    Reference compound is first; each additional compound appended with its ratio.
    Ratio = reference_conc / secondary_conc.
    """
    names  = list(compounds.keys())
    concs  = list(compounds.values())
    ref_conc = concs[0]
    ref_name = names[0]
    if len(names) == 1:
        return ref_name
    parts = [ref_name]
    for name, conc in zip(names[1:], concs[1:]):
        ratio = round(ref_conc / conc, 4)
        parts.append(f"{name} ({ratio})")
    return " / ".join(parts)


def generate_lhs(n: int, bounds: dict, seed: int) -> pd.DataFrame:
    """Generate LHS across 6 dimensions, returning g/L values for each row."""
    keys   = list(bounds.keys())
    lo     = [bounds[k][0] for k in keys]
    hi     = [bounds[k][1] for k in keys]
    sampler = qmc.LatinHypercube(d=6, seed=seed)
    sample  = sampler.random(n=n)
    scaled  = qmc.scale(sample, l_bounds=lo, u_bounds=hi)
    return pd.DataFrame(scaled, columns=keys)


# ── Main ──────────────────────────────────────────────────────────────────────

def generate_matrix(output_path: str = "media_matrix.xlsx"):

    # Generate LHS
    df_lhs = generate_lhs(N_CONDITIONS, BOUNDS, SEED)

    # Build row definitions: (compound_name_string, delft_conc, lhs_values)
    rows = []

    # Macronutrients — single compound per row
    for name, delft_conc in MACRONUTRIENTS.items():
        rows.append({
            "compound": name,
            "shorthand": name[0].upper(),
            "delft": delft_conc,
            "lhs": df_lhs[name].values,
        })

    # Trace metals — multi-compound row, reference is EDTA
    tm_name   = build_compound_name(TRACE_METALS)
    tm_delft  = list(TRACE_METALS.values())[0]   # EDTA concentration
    rows.append({
        "compound": tm_name,
        "shorthand": "TM",
        "delft": tm_delft,
        "lhs": df_lhs["Trace metals"].values,
    })

    # Vitamins — multi-compound row, reference is myo-Inositol
    vit_name  = build_compound_name(VITAMINS)
    vit_delft = list(VITAMINS.values())[0]        # myo-Inositol concentration
    rows.append({
        "compound": vit_name,
        "shorthand": "VIT",
        "delft": vit_delft,
        "lhs": df_lhs["Vitamins"].values,
    })

    # ── Build Main sheet dataframe ────────────────────────────────────────────

    condition_cols = ["1_Control"] + list(range(2, N_CONDITIONS + 2))
    main_cols = ["Compound", "ShorthandName", "PubChemCID"] + condition_cols

    shorthands = list("ABCDEF")
    main_rows = []
    for idx, r in enumerate(rows):
        row_data = {
            "Compound":      r["compound"],
            "ShorthandName": shorthands[idx],
            "PubChemCID":    None,   # not fetched automatically — fill manually if needed
            "1_Control":     r["delft"],
        }
        for i, val in enumerate(r["lhs"]):
            row_data[i + 2] = val
        main_rows.append(row_data)

    df_main = pd.DataFrame(main_rows, columns=main_cols)

    # ── Build Constraints sheet ───────────────────────────────────────────────

    constraints_rows = []
    for r in rows:
        constraints_rows.append({
            "Compound":                  r["compound"],
            "Lower Pipette Volume (μL)": "",   # to be filled by user/sb
            "Upper Pipette Volume (μL)": "",
            "Max Concentration (g/L)":   "",   # to be filled from solubility
        })
    df_constraints = pd.DataFrame(constraints_rows)

    # ── Build Miscellaneous sheet ─────────────────────────────────────────────

    df_misc = pd.DataFrame({
        "Max Volume (μL)":           [TOTAL_VOL_UL],
        "Maximum Algorithm Iterations": [MAX_ITER],
    })

    # ── Write to Excel ────────────────────────────────────────────────────────
    # Formatting matches reference file: Aptos Narrow 11pt, no bold, no fill, no freeze

    wb = openpyxl.Workbook()
    base_font = Font(name="Aptos Narrow", size=11)

    # # Custom formatting removed to match ProgrammaticPipetting.jl reference format:
    # # header_fill = PatternFill("solid", fgColor="2E5FA3")
    # # header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    # # ws_main.freeze_panes = "C2"
    # # ws_main.column_dimensions["A"].width = 60

    def write_sheet(ws, df):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = base_font
        for row_idx, row in df.iterrows():
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
                cell.font = base_font

    ws_main = wb.active
    ws_main.title = "Main"
    write_sheet(ws_main, df_main)

    ws_con = wb.create_sheet("Constraints")
    write_sheet(ws_con, df_constraints)

    ws_misc = wb.create_sheet("Miscellaneous")
    write_sheet(ws_misc, df_misc)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"  Main sheet: {len(rows)} compound rows x {len(condition_cols) + 2} columns")
    print(f"  Conditions: 1 control + {N_CONDITIONS} LHS")
    print()
    print("Compound rows:")
    for r in rows:
        print(f"  {r['shorthand']}: {r['compound'][:80]}...")
        print(f"    Delft (1x): {r['delft']} g/L")
        print(f"    LHS range:  {r['lhs'].min():.5f} — {r['lhs'].max():.5f} g/L")


if __name__ == "__main__":
    generate_matrix("media_matrix.xlsx")